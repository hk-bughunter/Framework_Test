from openpyxl import load_workbook
import os


class Read_data:

    def __init__(self, sheet=None):
        path = os.path.abspath("../data") + "/data.xlsx"

        if not sheet:
            sheet = "Sheet1"

        # 打开excel, 获取指定列表数据
        data = load_workbook(path)
        self.table = data[sheet]

    def check_data(self, row):
        row_list = []
        for i in self.table[row]:
            row_list.append(i.value)
        return row_list

    def get_data(self):
        data_list = []

        title = ",".join(self.check_data(1))
        data_list.append(title)

        value_list = []
        for row in range(2, self.table.max_row+1):
            values = tuple(self.check_data(row))
            value_list.append(values)

        data_list.append(value_list)

        return data_list







